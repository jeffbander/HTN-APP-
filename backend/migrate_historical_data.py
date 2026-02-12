"""
HTN-APP Historical Data Migration Script
=========================================
Imports data from AppSheet, Google Sheets enrollment, Microsoft Forms enrollment,
and Lifestyle Questionnaire into the new PostgreSQL database.

Run from backend/: python migrate_historical_data.py [--dry-run]

Data sources (place in backend/migration_data/):
  - AppSheet_ViewData_2026-02-09.csv              (user profiles)
  - AppSheet_ViewData_2026-02-09__1_.csv           (BP readings)
  - AppSheet_ViewData_2026-02-09__3_.csv           (call records)
  - Enrollment_Form_...1-492_.xlsx                 (MS Forms enrollment)
  - MSW HTN Prevention Program Enrollment Form.csv (Google Sheets enrollment)
  - Lifestyle_Questionnaire_1-265_.xlsx            (lifestyle data)

Migration tiers:
  Tier 1 — "app" users: 672 AppSheet active users (enriched with enrollment + lifestyle data)
  Tier 2 — "enrollment_only" users: ~333 MS Forms enrollees who never activated the app
"""

import sys
import os
import csv
import json
import re
import logging
from datetime import datetime
from collections import defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))

from app import create_app, db
from app.models.user import User
from app.models.reading import BloodPressureReading
from app.models.call_list_item import CallListItem
from app.models.call_attempt import CallAttempt
from app.models.union import Union

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DATA_DIR = os.path.join(os.path.dirname(__file__), 'migration_data')
DRY_RUN = '--dry-run' in sys.argv

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
log = logging.getLogger('migration')

# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
APPSHEET_USERS = os.path.join(DATA_DIR, 'AppSheet_ViewData_2026-02-09.csv')
APPSHEET_BP = os.path.join(DATA_DIR, 'AppSheet_ViewData_2026-02-09__1_.csv')
APPSHEET_CALLS = os.path.join(DATA_DIR, 'AppSheet_ViewData_2026-02-09__3_.csv')
MSFORMS_ENROLLMENT = os.path.join(DATA_DIR, 'Enrollment_Form_The_Mount_Sinai_Hypertension_Prevention_Program_for_First_Responders_1-492_.xlsx')
GSHEETS_ENROLLMENT = os.path.join(DATA_DIR, 'MSW HTN Prevention Program Enrollment Form.csv')
LIFESTYLE_Q = os.path.join(DATA_DIR, 'Lifestyle_Questionnaire_1-265_.xlsx')

# ---------------------------------------------------------------------------
# Union name → ID mapping (must match seed.py)
# ---------------------------------------------------------------------------
UNION_MAP = {
    'ufoa': 1,
    'ufa': 2,
    'ufadba': 3,
    'lba': 4,
    'mount sinai': 5,
    'other': 6,
}

def resolve_union(raw):
    """Map raw union string to union ID."""
    if not raw:
        return None
    return UNION_MAP.get(raw.strip().lower(), 6)  # default to 'Other'


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------
def clean(val):
    """Strip whitespace, return None for empty/None."""
    if val is None:
        return None
    v = str(val).strip()
    return v if v else None

def parse_bool(val):
    if val is None:
        return None
    v = str(val).strip().lower()
    if v in ('yes', 'true', '1', 'y'):
        return True
    if v in ('no', 'false', '0', 'n'):
        return False
    return None

def parse_int(val):
    if val is None:
        return None
    v = str(val).strip()
    m = re.search(r'(\d+)', v)
    return int(m.group(1)) if m else None

def parse_date(val, fallback_formats=None):
    """Try multiple date formats, return datetime or None."""
    if val is None:
        return None
    v = str(val).strip()
    if not v:
        return None
    formats = fallback_formats or [
        '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y',
        '%Y/%m/%d %I:%M:%S %p %Z', '%Y/%m/%d %H:%M:%S',
        '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %I:%M:%S %p',
        '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(v, fmt)
        except (ValueError, TypeError):
            continue
    # Try pandas-style as last resort
    try:
        from dateutil import parser as dateutil_parser
        return dateutil_parser.parse(v)
    except Exception:
        return None

def parse_height_inches(val):
    """Parse height strings like 5'10\", 5ft 10in, 70 inches, 5-10, etc."""
    if not val:
        return None
    v = str(val).strip().lower()
    # Pattern: 5'10" or 5' 10"
    m = re.search(r"(\d+)['\u2019]\s*(\d+)", v)
    if m:
        return int(m.group(1)) * 12 + int(m.group(2))
    # Pattern: 5ft 10in or 5 ft 10 in
    m = re.search(r'(\d+)\s*ft\s*(\d+)', v)
    if m:
        return int(m.group(1)) * 12 + int(m.group(2))
    # Pattern: 5-10 (feet-inches)
    m = re.search(r'^(\d)[- ](\d{1,2})$', v)
    if m:
        return int(m.group(1)) * 12 + int(m.group(2))
    # Just a number > 48 — assume inches
    m = re.search(r'^(\d{2,3})$', v)
    if m and int(m.group(1)) >= 48:
        return int(m.group(1))
    return None

def parse_weight_lbs(val):
    """Parse weight strings like '180 lbs', '180', etc."""
    if not val:
        return None
    v = str(val).strip().lower()
    if 'no weight' in v or v == 'none':
        return None
    m = re.search(r'(\d{2,3})', v)
    if m:
        w = int(m.group(1))
        if 70 <= w <= 500:
            return w
    return None

def parse_height_weight_combined(val):
    """Parse combined field like '5\\'10\" 180lbs' or '70 inches, 200 lbs'."""
    if not val:
        return None, None
    v = str(val).strip()
    height = parse_height_inches(v)
    weight = parse_weight_lbs(v)
    return height, weight

def build_food_frequency_json(food_dict):
    """Build JSON object from food frequency data. Returns JSON string or None."""
    if not food_dict or all(v is None for v in food_dict.values()):
        return None
    cleaned = {k: v for k, v in food_dict.items() if v is not None}
    return json.dumps(cleaned) if cleaned else None


# ---------------------------------------------------------------------------
# STEP 1: Load all data sources into memory
# ---------------------------------------------------------------------------
def load_appsheet_users():
    """Load AppSheet user profiles → dict keyed by lowercase email."""
    users = {}
    with open(APPSHEET_USERS, 'r') as f:
        reader = csv.reader(f)
        headers = next(reader)
        for row in reader:
            if len(row) < 2:
                continue
            email = row[1].strip().lower()
            if not email or '@' not in email:
                continue
            users[email] = row
    log.info(f'Loaded {len(users)} AppSheet user profiles')
    return users, headers

def load_appsheet_bp():
    """Load BP readings → list of dicts."""
    readings = []
    with open(APPSHEET_BP, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            email = row.get('Decode_Email', '').strip().lower()
            sbp = parse_int(row.get('SBP'))
            dbp = parse_int(row.get('DBP'))
            if not email or sbp is None or dbp is None:
                continue
            # Validate ranges
            if not (60 <= sbp <= 300 and 30 <= dbp <= 200):
                continue
            # Validate date range
            reading_dt = parse_date(row.get('Timestamp'))
            if reading_dt and (reading_dt < datetime(2020, 1, 1) or reading_dt > datetime(2026, 2, 9, 23, 59, 59)):
                continue
            readings.append({
                'email': email,
                'systolic': sbp,
                'diastolic': dbp,
                'heart_rate': parse_int(row.get('HR')),
                'reading_date': reading_dt,
                'created_at': parse_date(row.get('CreationDateTime')) or parse_date(row.get('Entry_Date')),
                'device_id': clean(row.get('Device ID')),
            })
    log.info(f'Loaded {len(readings)} valid BP readings')
    return readings

def load_appsheet_calls():
    """Load call records → list of dicts."""
    calls = []
    with open(APPSHEET_CALLS, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            patient = clean(row.get('Patient'))
            if not patient:
                continue
            calls.append({
                'patient_name': patient,
                'caller': clean(row.get('Caller')) or clean(row.get('Call_By')),
                'target_date': parse_date(row.get('Target_Date')),
                'date_of_call': parse_date(row.get('Date_of_Call')),
                'status': clean(row.get('Status')),
                'ai_analysis': clean(row.get('AI_Analysis')),
                'notes': clean(row.get('Notes')),
                'urls_to_send': clean(row.get('URLs_To_Send')),
                'documents_to_send': clean(row.get('Documents_To_Send')),
                'union': clean(row.get('Union')),
            })
    log.info(f'Loaded {len(calls)} call records')
    return calls

def load_msforms_enrollment():
    """Load MS Forms enrollment → dict keyed by lowercase email (from Email2 col)."""
    users = {}
    wb = openpyxl.load_workbook(MSFORMS_ENROLLMENT)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        email2 = str(row[10]).strip().lower() if row[10] else ''
        if not email2 or '@' not in email2:
            continue
        users[email2] = {
            'first_name': clean(row[6]),
            'last_name': clean(row[7]),
            'dob': clean(row[8]),
            'gender': clean(row[9]),
            'address': clean(row[11]),
            'phone': clean(row[12]),
            'union': clean(row[14]),
            'work_status': clean(row[15]),
            'rank': clean(row[16]),
            'chronic_conditions': clean(row[17]),
            'other_conditions': clean(row[18]),
            'has_hbp': clean(row[19]),
            'medications': clean(row[20]),
            'height_weight': clean(row[21]),
            'smoker': clean(row[22]),
            'race': clean(row[23]),
            'ethnicity': clean(row[24]),
            'start_time': row[1],  # form submission time
        }
    log.info(f'Loaded {len(users)} MS Forms enrollment records')
    return users

def load_gsheets_enrollment():
    """Load Google Sheets enrollment → dict keyed by lowercase email."""
    users = {}
    with open(GSHEETS_ENROLLMENT, 'r') as f:
        reader = csv.reader(f)
        next(reader)  # skip headers
        for row in reader:
            if len(row) < 10:
                continue
            email = row[5].strip().lower()
            if not email or '@' not in email:
                continue
            users[email] = {
                'first_name': clean(row[1]),
                'last_name': clean(row[2]),
                'dob': clean(row[3]),
                'gender': clean(row[4]),
                'address': clean(row[6]),
                'phone': clean(row[7]),
                'union': clean(row[9]),
                'work_status': clean(row[10]),
                'rank': clean(row[11]),
                'chronic_conditions': clean(row[12]),
                'other_conditions': clean(row[13]),
                'has_hbp': clean(row[14]),
                'medications': clean(row[15]),
                'height_weight': clean(row[16]),
                'smoker': clean(row[17]),
                'race': clean(row[18]),
                'ethnicity': clean(row[19]),
                'timestamp': clean(row[0]),
            }
    log.info(f'Loaded {len(users)} Google Sheets enrollment records')
    return users

def load_lifestyle_questionnaire():
    """Load Lifestyle Questionnaire → dict keyed by normalized 'first|last' name.
    For duplicates, keep the most recent submission."""
    entries = {}
    wb = openpyxl.load_workbook(LIFESTYLE_Q)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        first = str(row[6]).strip().lower() if row[6] else ''
        last = str(row[7]).strip().lower() if row[7] else ''
        if not first or not last:
            continue

        name_key = f"{re.sub(r'[^a-z]', '', first)}|{re.sub(r'[^a-z]', '', last)}"
        completion_time = row[2]  # Completion time for dedup

        # If duplicate, keep the most recent
        if name_key in entries:
            existing_time = entries[name_key].get('_completion_time')
            if existing_time and completion_time and completion_time <= existing_time:
                continue

        # Food frequency columns (indices 14-25)
        food_keys = [
            'fruit', 'vegetables', 'beans_nuts_seeds', 'fish_seafood',
            'whole_grains', 'refined_grains', 'low_fat_dairy',
            'high_fat_dairy', 'sweets', 'sweetened_beverages',
            'fried_foods', 'red_meat'
        ]
        food_data = {}
        for i, key in enumerate(food_keys):
            val = clean(row[14 + i])
            if val:
                food_data[key] = val

        entries[name_key] = {
            '_completion_time': completion_time,
            'on_bp_medication': parse_bool(row[10]),
            'missed_doses': parse_int(row[11]),
            'exercise_days': parse_int(row[12]),
            'exercise_minutes': parse_int(row[13]),
            'food_frequency': food_data,
            'financial_stress': clean(row[26]),
            'stress_level': clean(row[27]),
            'loneliness': clean(row[28]),
            'phq2_interest': clean(row[29]),
            'phq2_depressed': clean(row[30]),
            'sleep_quality': parse_int(row[31]),
        }
    log.info(f'Loaded {len(entries)} lifestyle questionnaire entries (deduplicated)')
    return entries


# ---------------------------------------------------------------------------
# STEP 2: Build enriched user profiles
# ---------------------------------------------------------------------------
def build_appsheet_user_profile(row, headers):
    """Parse a single AppSheet CSV row into a user dict."""
    def col(name):
        """Get value by approximate header match."""
        for i, h in enumerate(headers):
            if name.lower() in h.lower().strip("'\" "):
                return clean(row[i]) if i < len(row) else None
        return None

    # Parse food frequency from AppSheet (12 columns)
    food_keys = [
        ('fruit', 'Servings of Fruit'),
        ('vegetables', 'Servings of Vegetables'),
        ('beans_nuts_seeds', 'Servings of Beans'),
        ('fish_seafood', 'Servings of Fish'),
        ('whole_grains', 'Servings of Whole Grains'),
        ('refined_grains', 'Servings of refined'),
        ('low_fat_dairy', 'Servings of Low-fat'),
        ('high_fat_dairy', 'Servings of High-fat'),
        ('sweets', 'Servings of Sweets'),
        ('sweetened_beverages', 'Servings of Sweetened'),
        ('fried_foods', 'Servings of Fried'),
        ('red_meat', 'Servings of Red meat'),
    ]
    food_data = {}
    for key, header_fragment in food_keys:
        val = col(header_fragment)
        if val:
            food_data[key] = val

    name = col('Name')
    email = row[1].strip().lower() if len(row) > 1 else ''

    return {
        'name': name,
        'email': email,
        'dob': col('DOB'),
        'phone': col('Phone'),
        'address': col('Address'),
        'medications': col('Medications'),
        'gender': col('Gender'),
        'race': col('Race'),
        'ethnicity': col('Ethnicity'),
        'work_status': col('Work Status'),
        'union_raw': col('Union'),
        'height_raw': col('Height'),
        'weight_raw': col('Weight'),
        'chronic_conditions': col('PMH') or col('Chronic Conditions'),
        'has_hbp': col('HTN?'),
        'smoking_status': col('Smoking Status'),
        'exercise_days': col('Exericse Days') or col('Exercise Days'),
        'exercise_minutes': col('Exercise Minutes'),
        'food_frequency': food_data,
        'stress': col('Stress'),
        'sleep': col('Sleep'),
        'flag': col('Flag'),
        'notes': col('Notes'),
        'analysis': col('Analysis'),
    }

def gap_fill(profile, source, field_map):
    """Fill empty fields in profile from source using field_map.
    field_map: { profile_key: source_key }"""
    if not source:
        return
    for p_key, s_key in field_map.items():
        if not profile.get(p_key) and source.get(s_key):
            profile[p_key] = source[s_key]

def build_name_key(full_name):
    """Build normalized name key for lifestyle Q matching."""
    if not full_name:
        return None
    parts = full_name.strip().lower().split()
    if len(parts) < 2:
        return None
    first = re.sub(r'[^a-z]', '', parts[0])
    last = re.sub(r'[^a-z]', '', parts[-1])
    return f"{first}|{last}"


# ---------------------------------------------------------------------------
# STEP 3: Create database records
# ---------------------------------------------------------------------------
def create_user_record(profile, enrollment_source, lifestyle=None):
    """Create a User ORM object from enriched profile dict."""
    user = User()

    # PHI fields (encrypted via property setters)
    raw_name = profile.get('name') or f"{profile.get('first_name', '')} {profile.get('last_name', '')}".strip()
    user.name = raw_name or profile.get('email', '').split('@')[0] or 'Unknown'
    user.email = profile.get('email', '')
    user.dob = str(profile.get('dob', '')) if profile.get('dob') else None
    user.phone = profile.get('phone')
    user.address = profile.get('address')
    user.medications = profile.get('medications')

    # Demographics
    user.gender = profile.get('gender')
    user.race = profile.get('race')
    user.ethnicity = profile.get('ethnicity')
    user.work_status = profile.get('work_status')
    user.rank = profile.get('rank')
    user.smoking_status = profile.get('smoking_status') or profile.get('smoker')
    user.has_high_blood_pressure = parse_bool(profile.get('has_hbp'))

    # Height / weight
    if profile.get('height_raw'):
        user.height_inches = parse_height_inches(profile['height_raw'])
    if profile.get('weight_raw'):
        user.weight_lbs = parse_weight_lbs(profile['weight_raw'])
    # Gap-fill from combined field
    if profile.get('height_weight') and (not user.height_inches or not user.weight_lbs):
        h, w = parse_height_weight_combined(profile['height_weight'])
        if not user.height_inches and h:
            user.height_inches = h
        if not user.weight_lbs and w:
            user.weight_lbs = w

    # Chronic conditions → JSON array
    cc = profile.get('chronic_conditions')
    if cc:
        items = [c.strip() for c in re.split(r'[;,]', cc) if c.strip()]
        user.chronic_conditions = json.dumps(items) if items else None

    # Union
    user.union_id = resolve_union(profile.get('union_raw') or profile.get('union'))

    # Lifestyle data (from profile or Lifestyle Q overlay)
    if lifestyle:
        user.on_bp_medication = lifestyle.get('on_bp_medication')
        user.missed_doses = lifestyle.get('missed_doses')
        user.exercise_days_per_week = lifestyle.get('exercise_days') or parse_int(profile.get('exercise_days'))
        user.exercise_minutes_per_session = lifestyle.get('exercise_minutes') or parse_int(profile.get('exercise_minutes'))
        user.financial_stress = lifestyle.get('financial_stress')
        user.stress_level = lifestyle.get('stress_level') or profile.get('stress')
        user.loneliness = lifestyle.get('loneliness')
        user.sleep_quality = lifestyle.get('sleep_quality') or parse_int(profile.get('sleep'))
        user.phq2_interest = lifestyle.get('phq2_interest')
        user.phq2_depressed = lifestyle.get('phq2_depressed')

        # Food frequency: prefer Lifestyle Q, fallback to AppSheet
        lq_food = lifestyle.get('food_frequency', {})
        as_food = profile.get('food_frequency', {})
        merged_food = {**as_food, **lq_food}  # LQ overwrites AppSheet
        user.food_frequency = build_food_frequency_json(merged_food)
    else:
        user.exercise_days_per_week = parse_int(profile.get('exercise_days'))
        user.exercise_minutes_per_session = parse_int(profile.get('exercise_minutes'))
        user.stress_level = profile.get('stress')
        user.sleep_quality = parse_int(profile.get('sleep'))
        user.food_frequency = build_food_frequency_json(profile.get('food_frequency', {}))

    # System fields
    user.is_active = True
    user.is_approved = True
    user.is_flagged = parse_bool(profile.get('flag')) or False
    user.enrollment_source = enrollment_source

    return user


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def run_migration():
    log.info('=' * 60)
    log.info('HTN-APP Historical Data Migration')
    log.info(f'Dry run: {DRY_RUN}')
    log.info('=' * 60)

    # Check files exist
    for path in [APPSHEET_USERS, APPSHEET_BP, APPSHEET_CALLS,
                 MSFORMS_ENROLLMENT, GSHEETS_ENROLLMENT, LIFESTYLE_Q]:
        if not os.path.exists(path):
            log.error(f'Missing file: {path}')
            log.error(f'Place all data files in {DATA_DIR}/')
            sys.exit(1)

    # Load all sources
    log.info('\n--- Loading data sources ---')
    appsheet_users, appsheet_headers = load_appsheet_users()
    bp_readings = load_appsheet_bp()
    call_records = load_appsheet_calls()
    msforms = load_msforms_enrollment()
    gsheets = load_gsheets_enrollment()
    lifestyle = load_lifestyle_questionnaire()

    # Identify active user emails (AppSheet profiles + BP-reading-only)
    bp_emails = {r['email'] for r in bp_readings}
    active_emails = set(appsheet_users.keys()) | bp_emails

    # Gap-fill field mapping: enrollment source → profile field
    ENROLLMENT_FIELD_MAP = {
        'phone': 'phone',
        'address': 'address',
        'medications': 'medications',
        'rank': 'rank',
        'work_status': 'work_status',
        'gender': 'gender',
        'race': 'race',
        'ethnicity': 'ethnicity',
        'has_hbp': 'has_hbp',
        'smoking_status': 'smoker',
    }

    app = create_app()
    with app.app_context():
        # Ensure unions are seeded
        if Union.query.count() == 0:
            log.error('Unions table is empty. Run seed.py first.')
            sys.exit(1)

        # ---------------------------------------------------------------
        # TIER 1: Active "app" users (AppSheet profiles + BP-only users)
        # ---------------------------------------------------------------
        log.info('\n--- TIER 1: Creating active "app" users ---')
        email_to_user = {}  # email → User ORM object
        name_to_email = {}  # normalized name key → email (for call record matching)

        tier1_count = 0
        lifestyle_matched = 0

        for email in sorted(active_emails):
            # Start with AppSheet profile if it exists
            if email in appsheet_users:
                profile = build_appsheet_user_profile(
                    appsheet_users[email], appsheet_headers
                )
            else:
                # BP-only user — minimal profile
                profile = {'email': email, 'name': None}

            # Gap-fill from MS Forms first (priority 2)
            if email in msforms:
                gap_fill(profile, msforms[email], ENROLLMENT_FIELD_MAP)
                if not profile.get('dob') and msforms[email].get('dob'):
                    profile['dob'] = msforms[email]['dob']
                if not profile.get('name'):
                    first = msforms[email].get('first_name', '')
                    last = msforms[email].get('last_name', '')
                    profile['name'] = f"{first} {last}".strip()
                if not profile.get('union_raw') and msforms[email].get('union'):
                    profile['union_raw'] = msforms[email]['union']
                if msforms[email].get('height_weight'):
                    profile.setdefault('height_weight', msforms[email]['height_weight'])

            # Gap-fill from Google Sheets (priority 3)
            if email in gsheets:
                gap_fill(profile, gsheets[email], ENROLLMENT_FIELD_MAP)
                if not profile.get('dob') and gsheets[email].get('dob'):
                    profile['dob'] = gsheets[email]['dob']
                if not profile.get('name'):
                    first = gsheets[email].get('first_name', '')
                    last = gsheets[email].get('last_name', '')
                    profile['name'] = f"{first} {last}".strip()
                if not profile.get('union_raw') and gsheets[email].get('union'):
                    profile['union_raw'] = gsheets[email]['union']
                if gsheets[email].get('height_weight'):
                    profile.setdefault('height_weight', gsheets[email]['height_weight'])

            # Match Lifestyle Questionnaire by name
            nk = build_name_key(profile.get('name'))
            lq = lifestyle.get(nk) if nk else None
            if lq:
                lifestyle_matched += 1

            # Create user record
            user = create_user_record(profile, enrollment_source='app', lifestyle=lq)

            if not DRY_RUN:
                db.session.add(user)
                db.session.flush()  # get user.id

            email_to_user[email] = user
            if nk:
                name_to_email[nk] = email
            tier1_count += 1

        log.info(f'Tier 1 users created: {tier1_count}')
        log.info(f'Lifestyle Q matched: {lifestyle_matched}')

        # ---------------------------------------------------------------
        # TIER 2: MS Forms "enrollment_only" users
        # ---------------------------------------------------------------
        log.info('\n--- TIER 2: Creating "enrollment_only" users ---')
        tier2_count = 0
        for email, data in sorted(msforms.items()):
            if email in active_emails:
                continue  # already created in Tier 1

            profile = {
                'email': email,
                'name': f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
                'dob': data.get('dob'),
                'phone': data.get('phone'),
                'address': data.get('address'),
                'medications': data.get('medications'),
                'gender': data.get('gender'),
                'race': data.get('race'),
                'ethnicity': data.get('ethnicity'),
                'work_status': data.get('work_status'),
                'rank': data.get('rank'),
                'has_hbp': data.get('has_hbp'),
                'smoking_status': data.get('smoker'),
                'union_raw': data.get('union'),
                'height_weight': data.get('height_weight'),
                'chronic_conditions': data.get('chronic_conditions'),
            }

            # Also gap-fill from Google Sheets if they appear there
            if email in gsheets:
                gap_fill(profile, gsheets[email], ENROLLMENT_FIELD_MAP)

            user = create_user_record(profile, enrollment_source='enrollment_only')

            if not DRY_RUN:
                db.session.add(user)
                db.session.flush()

            email_to_user[email] = user
            nk = build_name_key(profile['name'])
            if nk:
                name_to_email[nk] = email
            tier2_count += 1

        log.info(f'Tier 2 users created: {tier2_count}')

        if not DRY_RUN:
            db.session.commit()
            log.info(f'Total users committed: {tier1_count + tier2_count}')

        # ---------------------------------------------------------------
        # IMPORT BP READINGS
        # ---------------------------------------------------------------
        log.info('\n--- Importing BP readings ---')
        bp_imported = 0
        bp_skipped = 0

        for r in bp_readings:
            user = email_to_user.get(r['email'])
            if not user:
                bp_skipped += 1
                continue

            reading = BloodPressureReading(
                user_id=user.id if not DRY_RUN else 0,
                systolic=r['systolic'],
                diastolic=r['diastolic'],
                heart_rate=r['heart_rate'],
                reading_date=r['reading_date'] or datetime.utcnow(),
                created_at=r['created_at'] or datetime.utcnow(),
                device_id=r['device_id'],
            )

            if not DRY_RUN:
                db.session.add(reading)
            bp_imported += 1

            # Batch commit every 1000
            if not DRY_RUN and bp_imported % 1000 == 0:
                db.session.commit()
                log.info(f'  ... {bp_imported} readings committed')

        if not DRY_RUN:
            db.session.commit()
        log.info(f'BP readings imported: {bp_imported}, skipped: {bp_skipped}')

        # ---------------------------------------------------------------
        # IMPORT CALL RECORDS
        # ---------------------------------------------------------------
        log.info('\n--- Importing call records ---')
        calls_imported = 0
        calls_skipped = 0

        # Build a name→user lookup for call matching
        name_to_user = {}
        for email, user in email_to_user.items():
            if user.name:
                nk = build_name_key(user.name)
                if nk:
                    name_to_user[nk] = user

        # Get or create system admin for call attempts
        if not DRY_RUN:
            system_admin = User.find_by_email('admin@bp-app.local')
            admin_id = system_admin.id if system_admin else 1
        else:
            admin_id = 0

        for call in call_records:
            # Match patient by name
            patient_nk = build_name_key(call['patient_name'])
            patient = name_to_user.get(patient_nk) if patient_nk else None

            if not patient:
                calls_skipped += 1
                continue

            # Create CallListItem
            cli = CallListItem(
                user_id=patient.id if not DRY_RUN else 0,
                list_type='coach',
                status='closed' if call['date_of_call'] else 'open',
                close_reason='resolved' if call['date_of_call'] else None,
                priority='medium',
                created_at=call['target_date'] or datetime.utcnow(),
                closed_at=call['date_of_call'],
            )

            if not DRY_RUN:
                db.session.add(cli)
                db.session.flush()

            # Create CallAttempt if there was an actual call
            if call['date_of_call']:
                # Map status to outcome
                status_raw = (call['status'] or '').lower()
                if 'complet' in status_raw or 'done' in status_raw:
                    outcome = 'completed'
                elif 'vm' in status_raw or 'voicemail' in status_raw:
                    outcome = 'left_vm'
                elif 'no answer' in status_raw or 'no_answer' in status_raw:
                    outcome = 'no_answer'
                elif 'email' in status_raw:
                    outcome = 'email_sent'
                else:
                    outcome = 'completed'  # default for historical

                # Combine notes
                notes_parts = []
                if call['notes']:
                    notes_parts.append(call['notes'])
                if call['ai_analysis']:
                    notes_parts.append(f"[AI Analysis] {call['ai_analysis']}")
                combined_notes = '\n\n'.join(notes_parts) if notes_parts else None

                # Check if materials were sent
                has_materials = bool(call['urls_to_send'] or call['documents_to_send'])
                materials_desc_parts = []
                if call['documents_to_send']:
                    materials_desc_parts.append(call['documents_to_send'])
                if call['urls_to_send']:
                    materials_desc_parts.append(call['urls_to_send'])

                attempt = CallAttempt(
                    call_list_item_id=cli.id if not DRY_RUN else 0,
                    user_id=patient.id if not DRY_RUN else 0,
                    admin_id=admin_id,
                    outcome=outcome,
                    materials_sent=has_materials,
                    materials_desc='; '.join(materials_desc_parts) if materials_desc_parts else None,
                    created_at=call['date_of_call'],
                )
                # Set encrypted notes via property
                if combined_notes:
                    attempt.notes = combined_notes

                if not DRY_RUN:
                    db.session.add(attempt)

            calls_imported += 1

        if not DRY_RUN:
            db.session.commit()
        log.info(f'Call records imported: {calls_imported}, skipped: {calls_skipped}')

        # ---------------------------------------------------------------
        # SUMMARY
        # ---------------------------------------------------------------
        log.info('\n' + '=' * 60)
        log.info('MIGRATION SUMMARY')
        log.info('=' * 60)
        log.info(f'Tier 1 users ("app"):            {tier1_count}')
        log.info(f'  ↳ with Lifestyle Q data:       {lifestyle_matched}')
        log.info(f'Tier 2 users ("enrollment_only"): {tier2_count}')
        log.info(f'Total users:                      {tier1_count + tier2_count}')
        log.info(f'BP readings imported:              {bp_imported}')
        log.info(f'BP readings skipped:               {bp_skipped}')
        log.info(f'Call records imported:              {calls_imported}')
        log.info(f'Call records skipped:               {calls_skipped}')
        if DRY_RUN:
            log.info('\n*** DRY RUN — no data was written to the database ***')
        log.info('=' * 60)


if __name__ == '__main__':
    run_migration()
